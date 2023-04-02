import boto3
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from dotenv import load_dotenv

from db.selections import *

load_dotenv()

# function to manage all the parameters of an aws service
def parameters_processing(params_df, instance):
    info = []
    # splitting the parameters inside the df
    for index, row in params_df.iterrows():
        param_name = row['name']
        param_name = param_name.split(';') # for parameters which are chained together with ;
        param_description = row['description']
        param_value = instance
        for i in range(len(param_name)):
            param_value = param_value[param_name[i]]
        # sending the information to the user (console + return function)
        print(f"{param_description}: {param_value}")
        info.append(f"{param_description}: {param_value}")
    return info

# function to manage all the services in a region and export them to excel
def services_management(region):
    # open the excel workbook
    workbook = load_workbook(filename=os.environ["workbook_name"])

    # getting all the services from the db in a df
    df_services = get_services()

    # adding a region sheet
    worksheet = workbook.create_sheet(title=region)
    sheet_row = 1

    # getting the information for each service in the df
    for service_id in range(1, df_services['id'].count()+1):
        service_df = df_services[df_services['id'] == service_id]
        service_name = service_df['name'].iloc[0]
        service_client = service_df['client'].iloc[0]
        service_describe_method = service_df['describe_method'].iloc[0]
        service_main_parameter = service_df['main_parameter'].iloc[0]
        service_secondary_parameter = service_df['secondary_parameter'].iloc[0]

        # getting al the parameters of the relevant service by the db
        parameters_df = get_parameters(service_id)

        # setting headers for the excel table
        service_headers = []
        for index, row in parameters_df.iterrows():
            service_headers.append(str(row['description']))

        Title = [f"This is a table for {service_name}"]
        worksheet.append(Title)
        staring_row = sheet_row + 1
        # white_space = [""]
        # worksheet.append(white_space)
        worksheet.append(service_headers)
        sheet_row += 2
        

        # connecting to aws api using boto3
        client = boto3.client(service_client, region_name=region)
        # client.getattr()()
        response = eval(f"client.{service_describe_method}()")

        
        # getting the instances of each service 
        for instance in response[service_main_parameter]:
            if service_secondary_parameter != "":
                # meaning a group of instances, need to unpack them farther
                for sub_instance in instance[service_secondary_parameter]:
                    # getting the information for each parameter about the instance
                    service_info = parameters_processing(parameters_df, sub_instance)
                    # adding the information to the excel
                    worksheet.append(service_info)
            else:
                # getting the information for each parameter about the instance
                # print(type(parameters_processing(parameters_df, instance)))
                # print(type((parameters_processing(parameters_df, instance))))
                service_info = parameters_processing(parameters_df, instance)
                # adding the information to the excel
                worksheet.append(service_info)
            sheet_row += 1
            print("--------------------")
        # Define the table properties
        end_column = parameters_df['id'].count()
        ref = "{}{}:{}{}".format('A', staring_row, chr(64 + end_column), sheet_row)
        table = Table(displayName=f'{region}-{service_name}', ref=ref)

        # Set the style of the table
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                            showRowStripes=True, showColumnStripes=False)

        # Set the style to the table
        table.tableStyleInfo = style

        # Add the table to the worksheet
        worksheet.add_table(table)
        white_space = [""]
        worksheet.append(white_space)
        worksheet.append(white_space)
        sheet_row += 2

    # Close the workbook
    workbook.save(filename=os.environ["workbook_name"])

# function to get all the aws regions and apply the services_management function for each region
def regions_management():
    # getting al the regions
    regions = [region['RegionName'] for region in boto3.client('ec2').describe_regions()['Regions']]
    for region in regions:
        print("------------------")
        print(f"starting to scan {region}")
        print("------------------")
        # start the services scan
        services_management(region)


# initialize the app
regions_management()